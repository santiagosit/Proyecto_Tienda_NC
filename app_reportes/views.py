# Django imports
from django.shortcuts import render
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required, user_passes_test
from django.db.models import Sum, F, Q, Count
from django.utils import timezone
from django.db.models.functions import TruncMonth, TruncWeek
from django.template.defaultfilters import floatformat

# Third-party imports
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
from openpyxl.chart import BarChart, Reference, LineChart
from datetime import datetime, timedelta
import json
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from io import BytesIO
import matplotlib.pyplot as plt
import base64

# Local imports
from app_usuarios.utils import is_admin_or_superuser
from app_inventario.models import Producto
from app_ventas.models import Venta, VentaDetalle
from app_finanzas.models import Ingreso, Egreso
from app_pedidos.models import Pedido

# Helper functions
def estilizar_encabezado(celda):
    """Aplica estilos al encabezado de Excel."""
    celda.font = Font(bold=True, color="FFFFFF")
    celda.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    celda.alignment = Alignment(horizontal="center", vertical="center")
    celda.border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

def productos_mas_vendidos(tipo_tiempo='mensual'):
    """Retorna los productos más vendidos en un periodo."""
    hoy = timezone.now().date()
    inicio_periodo = (
        hoy - timezone.timedelta(days=7)
        if tipo_tiempo == 'semanal'
        else hoy.replace(day=1)
    )

    return VentaDetalle.objects.filter(
        venta__fecha__range=[inicio_periodo, hoy]
    ).values('producto__nombre').annotate(
        total_vendido=Sum('cantidad')
    ).order_by('-total_vendido')

def productos_sin_stock():
    """Retorna productos sin stock."""
    return Producto.objects.filter(cantidad_stock=0)

def productos_no_vendidos():
    """Retorna productos sin ventas en los últimos 30 días."""
    hace_30_dias = timezone.now().date() - timezone.timedelta(days=30)
    return Producto.objects.exclude(
        id__in=VentaDetalle.objects.filter(
            venta__fecha__gte=hace_30_dias
        ).values('producto')
    )

# View functions
@login_required
@user_passes_test(is_admin_or_superuser)
def reporte_inventario(request):
    context = {
        'productos_mas_vendidos': VentaDetalle.objects.values(
            'producto__nombre'
        ).annotate(
            total_vendido=Sum('cantidad')
        ).order_by('-total_vendido')[:10],
        'productos_sin_stock': productos_sin_stock(),
        'productos_agotandose': Producto.objects.filter(
            cantidad_stock__lt=F('stock_minimo')
        ).order_by('cantidad_stock'),
        'productos_no_vendidos': Producto.objects.exclude(
            ventadetalle__isnull=False
        ),
    }
    return render(request, 'reportes/reporte_inventario.html', context)

@login_required
@user_passes_test(is_admin_or_superuser)
def reporte_ingresos_egresos(request):
    tipo_tiempo = request.GET.get('tipo_tiempo', 'mensual')
    hoy = timezone.now()
    
    if tipo_tiempo == 'semanal':
        # Última semana
        inicio_periodo = hoy - timedelta(days=7)
    else:
        # Mes actual
        inicio_periodo = hoy.replace(day=1, hour=0, minute=0, second=0, microsecond=0)

    # Obtener ingresos
    ingresos_por_periodo = Ingreso.objects.filter(
        fecha__gte=inicio_periodo,
        fecha__lte=hoy
    ).values('fecha__date').annotate(
        total=Sum('monto')
    ).order_by('fecha__date')

    # Obtener egresos
    egresos_por_periodo = Egreso.objects.filter(
        fecha__gte=inicio_periodo,
        fecha__lte=hoy
    ).values('fecha__date').annotate(
        total=Sum('monto')
    ).order_by('fecha__date')

    # Crear diccionarios para datos
    ingresos_dict = {x['fecha__date']: float(x['total']) if x['total'] else 0 for x in ingresos_por_periodo}
    egresos_dict = {x['fecha__date']: float(x['total']) if x['total'] else 0 for x in egresos_por_periodo}

    # Generar lista de fechas
    fechas = []
    fecha_actual = inicio_periodo.date()
    while fecha_actual <= hoy.date():
        fechas.append(fecha_actual)
        fecha_actual += timedelta(days=1)

    # Preparar listas de datos
    ingresos_lista = [ingresos_dict.get(fecha, 0) for fecha in fechas]
    egresos_lista = [egresos_dict.get(fecha, 0) for fecha in fechas]
    
    # Calcular totales
    total_ingresos = sum(ingresos_lista)
    total_egresos = sum(egresos_lista)
    balance = total_ingresos - total_egresos

    # Preparar datos para los gráficos
    chart_data = {
        'line_chart': {
            'labels': [fecha.strftime('%d/%m/%Y') for fecha in fechas],
            'ingresos': ingresos_lista,
            'egresos': egresos_lista
        },
        'pie_chart': {
            'labels': ['Ingresos', 'Egresos'],
            'data': [total_ingresos, total_egresos]
        }
    }

    context = {
        'tipo_tiempo': tipo_tiempo,
        'total_ingresos': total_ingresos,
        'total_egresos': total_egresos,
        'balance': balance,
        'chart_data': json.dumps(chart_data),
        'inicio_periodo': inicio_periodo.strftime('%d/%m/%Y'),
        'fin_periodo': hoy.strftime('%d/%m/%Y')
    }
    
    return render(request, 'reportes/reporte_ingresos_egresos.html', context)

@login_required
@user_passes_test(is_admin_or_superuser)
def exportar_reporte_excel(request):
    hoy = timezone.now().date()

    # Creamos un nuevo libro de Excel
    wb = openpyxl.Workbook()

    # --- Hoja 1: Productos más vendidos ---
    ws1 = wb.active
    ws1.title = 'Productos más vendidos'

    # Encabezados
    ws1.append(['Producto', 'Total Vendido'])
    for celda in ws1[1]:
        estilizar_encabezado(celda)

    # Datos
    productos_mas_vendidos = VentaDetalle.objects.values('producto__nombre').annotate(
        total_vendido=Sum('cantidad')).order_by('-total_vendido')[:10]
    for producto in productos_mas_vendidos:
        ws1.append([producto['producto__nombre'], producto['total_vendido']])

    # Ajustar ancho de columnas
    for col in ws1.columns:
        max_length = max(len(str(cell.value)) for cell in col)
        ws1.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

    # Gráfico de barras para productos más vendidos
    chart1 = BarChart()
    data = Reference(ws1, min_col=2, min_row=1, max_row=len(productos_mas_vendidos) + 1)
    cats = Reference(ws1, min_col=1, min_row=2, max_row=len(productos_mas_vendidos) + 1)
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(cats)
    chart1.title = "Productos más vendidos"
    chart1.y_axis.title = "Cantidad vendida"
    chart1.x_axis.title = "Producto"
    ws1.add_chart(chart1, "E5")

    # --- Hoja 2: Productos sin stock ---
    ws2 = wb.create_sheet('Productos sin stock')

    # Encabezados
    ws2.append(['Producto'])
    for celda in ws2[1]:
        estilizar_encabezado(celda)

    # Datos
    productos_sin_stock = Producto.objects.filter(cantidad_stock=0)
    for producto in productos_sin_stock:
        ws2.append([producto.nombre])

    # Ajustar ancho de columnas
    for col in ws2.columns:
        max_length = max(len(str(cell.value)) for cell in col)
        ws2.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

    # --- Hoja 3: Productos que se agotan más rápido ---
    ws3 = wb.create_sheet('Productos agotándose')

    # Encabezados
    ws3.append(['Producto', 'Stock actual'])
    for celda in ws3[1]:
        estilizar_encabezado(celda)

    # Datos
    productos_agotandose = Producto.objects.filter(cantidad_stock__lt=F('stock_minimo')).order_by('cantidad_stock')
    for producto in productos_agotandose:
        ws3.append([producto.nombre, producto.cantidad_stock])

    # Gráfico de barras para productos que se agotan más rápido
    chart2 = BarChart()
    data = Reference(ws3, min_col=2, min_row=1, max_row=len(productos_agotandose) + 1)
    cats = Reference(ws3, min_col=1, min_row=2, max_row=len(productos_agotandose) + 1)
    chart2.add_data(data, titles_from_data=True)
    chart2.set_categories(cats)
    chart2.title = "Productos con bajo stock"
    chart2.y_axis.title = "Stock actual"
    chart2.x_axis.title = "Producto"
    ws3.add_chart(chart2, "E5")

    # --- Hoja 4: Productos que no se venden ---
    ws4 = wb.create_sheet('Productos no vendidos')

    # Encabezados
    ws4.append(['Producto'])
    for celda in ws4[1]:
        estilizar_encabezado(celda)

    # Datos
    productos_no_vendidos = Producto.objects.exclude(ventadetalle__isnull=False)
    for producto in productos_no_vendidos:
        ws4.append([producto.nombre])


    # Configurar la respuesta HTTP para la descarga de Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=reporte_inventario.xlsx'

    # Guardar el archivo en el response
    wb.save(response)
    return response

@login_required
@user_passes_test(is_admin_or_superuser)
def exportar_reporte_financiero(request):
    tipo_tiempo = request.GET.get('tipo_tiempo', 'mensual')
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte Financiero"

    # Estilos
    header_style = NamedStyle(name="header_style")
    header_style.font = Font(bold=True)
    header_style.fill = PatternFill("solid", fgColor="CCE5FF")

    # Encabezados
    headers = ['Fecha', 'Ingresos', 'Egresos', 'Balance']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.style = header_style

    # Datos
    hoy = timezone.now().date()
    inicio_periodo = hoy - timedelta(days=7) if tipo_tiempo == 'semanal' else hoy.replace(day=1)

    row = 2
    for fecha in (inicio_periodo + timedelta(n) for n in range((hoy - inicio_periodo).days + 1)):
        ingresos = Ingreso.objects.filter(fecha__date=fecha).aggregate(Sum('monto'))['monto__sum'] or 0
        egresos = Egreso.objects.filter(fecha__date=fecha).aggregate(Sum('monto'))['monto__sum'] or 0
        balance = ingresos - egresos

        ws.cell(row=row, column=1, value=fecha.strftime('%Y-%m-%d'))
        ws.cell(row=row, column=2, value=float(ingresos))
        ws.cell(row=row, column=3, value=float(egresos))
        ws.cell(row=row, column=4, value=float(balance))
        row += 1

    # Ajustar ancho de columnas
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2

    # Agregar gráfico
    chart = LineChart()
    chart.title = "Ingresos vs Egresos"
    chart.style = 12
    chart.height = 10
    chart.width = 20

    data = Reference(ws, min_col=2, min_row=1, max_row=row-1, max_col=3)
    cats = Reference(ws, min_col=1, min_row=2, max_row=row-1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    ws.add_chart(chart, "F2")

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=reporte_financiero_{tipo_tiempo}.xlsx'
    wb.save(response)
    return response

@login_required
@user_passes_test(is_admin_or_superuser)
def exportar_reporte_pdf(request):
    # Crear el objeto de respuesta HTTP
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="reporte_financiero.pdf"'

    # Crear el documento PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []

    # Estilos
    styles = getSampleStyleSheet()
    title_style = styles['Heading1']
    normal_style = styles['Normal']

    # Título
    elements.append(Paragraph('Reporte Financiero', title_style))
    elements.append(Spacer(1, 12))

    # Obtener datos financieros
    tipo_tiempo = request.GET.get('tipo_tiempo', 'mensual')
    hoy = timezone.now()
    
    if tipo_tiempo == 'semanal':
        inicio_periodo = hoy - timedelta(days=7)
    else:
        inicio_periodo = hoy.replace(day=1)

    # Obtener ingresos y egresos
    ingresos = Ingreso.objects.filter(
        fecha__gte=inicio_periodo,
        fecha__lte=hoy
    ).values('fecha__date').annotate(
        total=Sum('monto')
    ).order_by('fecha__date')

    egresos = Egreso.objects.filter(
        fecha__gte=inicio_periodo,
        fecha__lte=hoy
    ).values('fecha__date').annotate(
        total=Sum('monto')
    ).order_by('fecha__date')

    # Resumen financiero
    total_ingresos = sum(x['total'] for x in ingresos)
    total_egresos = sum(x['total'] for x in egresos)
    balance = total_ingresos - total_egresos

    # Tabla de resumen
    data = [
        ['Concepto', 'Monto'],
        ['Total Ingresos', f'${total_ingresos:,.2f}'],
        ['Total Egresos', f'${total_egresos:,.2f}'],
        ['Balance', f'${balance:,.2f}']
    ]

    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 14),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))

    elements.append(table)
    elements.append(Spacer(1, 20))

    # Crear lista completa de fechas
    fechas = []
    fecha_actual = inicio_periodo.date()
    while fecha_actual <= hoy.date():
        fechas.append(fecha_actual)
        fecha_actual += timedelta(days=1)

    # Crear diccionarios para ingresos y egresos
    ingresos_dict = {x['fecha__date']: float(x['total']) for x in ingresos}
    egresos_dict = {x['fecha__date']: float(x['total']) for x in egresos}

    # Crear listas alineadas
    montos_ingresos = [ingresos_dict.get(fecha, 0) for fecha in fechas]
    montos_egresos = [egresos_dict.get(fecha, 0) for fecha in fechas]

    # Generar gráficas
    plt.figure(figsize=(10, 6))
    plt.plot(fechas, montos_ingresos, 'g-', label='Ingresos', marker='o')
    plt.plot(fechas, montos_egresos, 'r-', label='Egresos', marker='o')
    plt.title('Ingresos vs Egresos')
    plt.xlabel('Fecha')
    plt.ylabel('Monto ($)')
    plt.legend()
    plt.grid(True)
    plt.xticks(rotation=45)
    plt.tight_layout()

    # Guardar gráfico
    img_buffer = BytesIO()
    plt.savefig(img_buffer, format='png', dpi=300, bbox_inches='tight')
    img_buffer.seek(0)
    plt.close()

    # Crear imagen para el PDF
    img = Image(img_buffer)
    img.drawHeight = 4*inch
    img.drawWidth = 7*inch
    elements.append(img)

    # Agregar tabla de datos
    data = [['Fecha', 'Ingresos', 'Egresos']]
    for fecha, ingreso, egreso in zip(fechas, montos_ingresos, montos_egresos):
        data.append([
            fecha.strftime('%d/%m/%Y'),
            f'${ingreso:,.2f}',
            f'${egreso:,.2f}'
        ])

    # Estilo de tabla
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
    ]))

    elements.append(Spacer(1, 20))
    elements.append(table)

    # Construir PDF
    doc.build(elements)
    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)
    
    return response

def home(request):
    context = {
        'productos_bajo_stock_count': Producto.objects.filter(
            cantidad_stock__lt=F('stock_minimo')
        ).count(),
        'productos_sin_stock_count': productos_sin_stock().count(),
        'pedidos_pendientes_count': Pedido.objects.filter(
            Q(estado='pedido') | Q(estado='en camino')
        ).count(),
    }
    return render(request, 'home.html', context)

